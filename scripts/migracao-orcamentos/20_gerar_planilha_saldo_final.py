import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
REPORT_DIR = ROOT / "outputs" / "migracao-orcamentos" / "saldo-final-2026-07-21"
DATA_PATH = REPORT_DIR / "saldo_final.json"
OUTPUT_PATH = REPORT_DIR / "orcamentos_saldo_final_sem_cancelados_2026-07-21.xlsx"

NAVY = "17365D"
BLUE = "1F4E78"
BLUE_LIGHT = "DCE6F1"
PALE = "F4F8FB"
WHITE = "FFFFFF"
TEXT = "1F2937"
MUTED = "64748B"
GRID = "CBD5E1"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F4CCCC"
PURPLE = "E4DFEC"
THIN = Side(style="thin", color=GRID)


def parse_date(value):
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value.replace(tzinfo=None) if isinstance(value, datetime) else value
    text = str(value).strip()
    for parser in (
        lambda: datetime.fromisoformat(text.replace("Z", "+00:00")),
        lambda: datetime.strptime(text[:10], "%Y-%m-%d"),
        lambda: datetime.strptime(text[:10], "%d/%m/%Y"),
    ):
        try:
            parsed = parser()
            return parsed.replace(tzinfo=None) if isinstance(parsed, datetime) else parsed
        except (ValueError, TypeError):
            pass
    return text


def title_block(ws, title, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=15)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=PALE)
    ws.cell(2, 1).font = Font(color=TEXT, italic=True, size=10)
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34
    ws.sheet_view.showGridLines = False


def style_header(ws, row, end_col):
    for column in range(1, end_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[row].height = 36


def apply_page_setup(ws, repeat_rows=None):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.oddFooter.center.text = "Ipromed - migração de orçamentos"
    ws.oddFooter.right.text = "Página &P de &N"
    if repeat_rows:
        ws.print_title_rows = repeat_rows


def set_widths(ws, widths):
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def add_table(ws, name, header_row, end_row, end_col):
    if end_row <= header_row:
        return
    reference = f"A{header_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def value_for_excel(value, header):
    if header in {"Data de criação", "Data de cancelamento"}:
        return parse_date(value)
    if header in {"Valor total", "Soma dos itens", "Diferença"}:
        return float(value or 0)
    return value


COMPACT_COLUMNS = [
    ("ID ArkMeds", "arkmeds_orcamento_id"),
    ("Número", "numero"),
    ("Solicitante", "solicitante"),
    ("Status atual", "status_arkmeds_atual"),
    ("Status da planilha", "status_planilha"),
    ("Data de criação", "data_criacao"),
    ("Data de cancelamento", "data_cancelamento"),
    ("Tipo de orçamento", "tipo_orcamento"),
    ("Valor total", "valor_total"),
    ("Soma dos itens", "soma_itens"),
    ("Diferença", "diferenca_valor"),
    ("Serviços", "quantidade_servicos"),
    ("Peças", "quantidade_pecas"),
    ("Itens no staging", "quantidade_itens_staging"),
    ("OS de origem", "ordem_servico_origem"),
    ("OS resolvida", "ordem_servico_resolvida"),
    ("Empresa resolvida", "empresa_resolvida"),
    ("Identificador", "identificador"),
    ("Equipamento", "equipamento"),
    ("Fabricante", "fabricante"),
    ("Modelo", "modelo"),
    ("N. série", "numero_serie"),
    ("Patrimônio", "patrimonio"),
    ("Vínculo com OS", "classificacao_vinculo_os"),
    ("Validação", "status_validacao"),
    ("Motivos bloqueantes", "motivos_bloqueantes"),
    ("Avisos", "avisos_validacao"),
    ("Itens preservados", "tem_itens_preservados"),
    ("Ação recomendada", "acao_recomendada"),
    ("Itens - resumo", "itens_resumo"),
    ("Observações da planilha", "observacoes_planilha"),
    ("PDF original", "pdf_original_url"),
]


def create_detail_sheet(wb, name, title, subtitle, rows, table_name):
    ws = wb.create_sheet(name)
    title_block(ws, title, subtitle, len(COMPACT_COLUMNS))
    header_row = 4
    for column, (header, _) in enumerate(COMPACT_COLUMNS, 1):
        ws.cell(header_row, column, header)
    style_header(ws, header_row, len(COMPACT_COLUMNS))
    ws.cell(header_row, 25).comment = Comment(
        "Status técnico do staging antes da importação definitiva.",
        "Codex",
    )
    ws.cell(header_row, 29).comment = Comment(
        "Orientação para o próximo passo; nenhuma importação foi executada por esta planilha.",
        "Codex",
    )

    wrap_columns = {3, 8, 17, 18, 19, 24, 26, 27, 29, 30, 31, 32}
    money_columns = {9, 10, 11}
    date_columns = {6, 7}
    for row_number, item in enumerate(rows, header_row + 1):
        for column, (header, key) in enumerate(COMPACT_COLUMNS, 1):
            cell = ws.cell(row_number, column, value_for_excel(item.get(key), header))
            cell.font = Font(color=TEXT, size=9)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column in wrap_columns,
            )
            if column in money_columns:
                cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
            if column in date_columns and isinstance(cell.value, (date, datetime)):
                cell.number_format = "dd/mm/yyyy"

        validation = str(item.get("status_validacao") or "")
        fill = GREEN if validation == "historico_consulta" else (
            ORANGE if validation == "pendente_os" else (
                YELLOW if validation == "pendente_itens" else RED
            )
        )
        ws.cell(row_number, 25).fill = PatternFill("solid", fgColor=fill)
        if item.get("acao_recomendada") == "Incluir no dry-run de cancelados":
            ws.cell(row_number, 29).fill = PatternFill("solid", fgColor=GREEN)

    widths = [12, 13, 34, 14, 17, 14, 16, 20, 14, 14, 14, 10, 10, 12, 14, 14,
              32, 38, 28, 20, 20, 18, 16, 22, 21, 32, 32, 14, 34, 60, 55, 46]
    set_widths(ws, widths)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(COMPACT_COLUMNS))}{max(header_row, ws.max_row)}"
    add_table(ws, table_name, header_row, ws.max_row, len(COMPACT_COLUMNS))
    apply_page_setup(ws, "$4:$4")
    return ws


def create_new_sheet(wb, rows):
    columns = [
        ("ID ArkMeds", "arkmeds_orcamento_id"),
        ("Número", "numero"),
        ("Solicitante", "solicitante"),
        ("Status atual", "status_arkmeds_atual"),
        ("Data de criação", "data_criacao"),
        ("Valor total", "valor_total"),
        ("Ação recomendada", "acao_recomendada"),
    ]
    ws = wb.create_sheet("Novos sem staging")
    title_block(
        ws,
        "Orçamentos novos ainda não coletados para o staging",
        "Estes registros existem no ArkMeds, mas ainda precisam da coleta de cabeçalho e itens antes de qualquer importação.",
        len(columns),
    )
    for column, (header, _) in enumerate(columns, 1):
        ws.cell(4, column, header)
    style_header(ws, 4, len(columns))
    for row_number, item in enumerate(rows, 5):
        for column, (header, key) in enumerate(columns, 1):
            value = item.get(key)
            if header == "Data de criação":
                value = parse_date(value)
            if header == "Valor total":
                value = float(value or 0)
            cell = ws.cell(row_number, column, value)
            cell.border = Border(bottom=THIN)
            cell.font = Font(color=TEXT, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=column in {3, 7})
            if header == "Data de criação" and isinstance(value, (date, datetime)):
                cell.number_format = "dd/mm/yyyy"
            if header == "Valor total":
                cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
        ws.cell(row_number, 7).fill = PatternFill("solid", fgColor=YELLOW)
    set_widths(ws, [14, 14, 42, 18, 16, 18, 44])
    ws.freeze_panes = "A5"
    add_table(ws, "TabelaNovosSemStaging", 4, ws.max_row, len(columns))
    apply_page_setup(ws, "$4:$4")


def create_original_data_sheet(wb, rows):
    source_headers = []
    for item in rows:
        for key in (item.get("dados_planilha") or {}).keys():
            if key not in source_headers:
                source_headers.append(key)

    diagnostic_columns = [
        ("Status ArkMeds atual", "status_arkmeds_atual"),
        ("Status validação staging", "status_validacao"),
        ("Motivos bloqueantes atuais", "motivos_bloqueantes"),
        ("OS resolvida Ipromed", "ordem_servico_resolvida"),
        ("Empresa resolvida Ipromed", "empresa_resolvida"),
        ("Ação recomendada", "acao_recomendada"),
    ]
    all_headers = source_headers + [header for header, _ in diagnostic_columns]
    ws = wb.create_sheet("Dados Originais")
    title_block(
        ws,
        "Dados originais dos orçamentos ainda não importados",
        "A estrutura original foi preservada. As seis últimas colunas foram acrescentadas para mostrar o diagnóstico atual da migração.",
        max(1, len(all_headers)),
    )
    for column, header in enumerate(all_headers, 1):
        ws.cell(4, column, header)
    style_header(ws, 4, len(all_headers))

    for row_number, item in enumerate(rows, 5):
        source = item.get("dados_planilha") or {}
        values = [source.get(header) for header in source_headers]
        values.extend(item.get(key) for _, key in diagnostic_columns)
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_number, column, value)
            cell.font = Font(color=TEXT, size=9)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, header in enumerate(all_headers, 1):
        width = 16
        lowered = str(header).lower()
        if any(term in lowered for term in ("solicitante", "observ", "descr", "motivo", "identifica", "ação", "acao")):
            width = 36
        elif any(term in lowered for term in ("data", "valor", "número", "numero", "id")):
            width = 15
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(all_headers))}{ws.max_row}"
    add_table(ws, "TabelaDadosOriginaisSaldo", 4, ws.max_row, len(all_headers))
    apply_page_setup(ws, "$4:$4")


def create_summary(wb, payload):
    summary = payload["resumo"]
    ws = wb.active
    ws.title = "Resumo"
    title_block(
        ws,
        "Saldo final da migração de orçamentos",
        "Posição atual em 21/07/2026, após aplicar descontos, vínculos exatos e remover registros vazios. O ArkMeds foi consultado somente para leitura.",
        8,
    )

    ws.merge_cells("A4:H4")
    ws["A4"] = "Situação atual"
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A4"].font = Font(color=WHITE, bold=True)
    ws["A4"].alignment = Alignment(vertical="center")

    metrics = [
        ("Total atual no ArkMeds", summary["total_arkmeds_atual"]),
        ("Já importados na Ipromed", summary["total_importado_ipromed"]),
        ("Restantes no staging", summary["total_restante_staging"]),
        ("Cancelados restantes", summary["cancelados_restantes"]),
        ("Outros restantes", summary["outros_restantes"]),
        ("Novos sem staging", summary["novos_sem_staging"]),
    ]
    for index, (label, value) in enumerate(metrics):
        row = 6 + index // 3 * 3
        column = 1 + index % 3 * 2
        ws.cell(row, column, label)
        ws.cell(row, column).font = Font(color=MUTED, size=9)
        ws.cell(row + 1, column, value)
        ws.cell(row + 1, column).font = Font(color=NAVY, bold=True, size=16)
        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        ws.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        fill = GREEN if label == "Já importados na Ipromed" else (
            ORANGE if label == "Cancelados restantes" else PALE
        )
        for current_row in (row, row + 1):
            for current_column in (column, column + 1):
                ws.cell(current_row, current_column).fill = PatternFill("solid", fgColor=fill)
                ws.cell(current_row, current_column).border = Border(bottom=THIN)

    ws["A13"] = "Cancelados restantes por validação"
    ws["A13"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A13"].font = Font(color=WHITE, bold=True)
    ws.merge_cells("A13:D13")
    ws["F13"] = "Importados por status"
    ws["F13"].fill = PatternFill("solid", fgColor=BLUE)
    ws["F13"].font = Font(color=WHITE, bold=True)
    ws.merge_cells("F13:H13")

    validation_labels = {
        "historico_consulta": "Histórico - iniciar dry-run",
        "pendente_os": "Pendente de OS/vínculo",
        "pendente_itens": "Pendente de itens",
        "pendente_valor": "Pendente de valor",
    }
    for index, (status, count) in enumerate(summary["cancelados_por_validacao"].items(), 14):
        ws.cell(index, 1, validation_labels.get(status, status))
        ws.cell(index, 4, count)
        ws.cell(index, 4).font = Font(bold=True)
        for column in range(1, 5):
            ws.cell(index, column).border = Border(bottom=THIN)

    imported_labels = {
        "pendente": "Pendentes",
        "aprovado": "Aprovados",
        "reprovado": "Reprovados",
        "faturado": "Faturados",
        "cancelado": "Cancelados",
    }
    for index, (status, count) in enumerate(summary["importados_por_status"].items(), 14):
        ws.cell(index, 6, imported_labels.get(status, status))
        ws.cell(index, 8, count)
        ws.cell(index, 8).font = Font(bold=True)
        for column in range(6, 9):
            ws.cell(index, column).border = Border(bottom=THIN)

    note_row = 21
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=8)
    historicos_restantes = summary["cancelados_por_validacao"].get("historico_consulta", 0)
    ws.cell(note_row, 1, (
        f"Saldo atualizado após aplicar descontos, vínculos exatos e retirar registros vazios. Restam "
        f"{summary['cancelados_restantes']} cancelados para revisão, dos quais {historicos_restantes} "
        f"estão classificados como histórico. Os {summary['novos_sem_staging']} registros novos "
        "precisam ser coletados para o staging antes de entrar em qualquer lote."
    ))
    ws.cell(note_row, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(note_row, 1).fill = PatternFill("solid", fgColor=YELLOW)
    ws.cell(note_row, 1).font = Font(color=TEXT, bold=True)
    set_widths(ws, [24, 16, 24, 16, 4, 24, 16, 16])
    apply_page_setup(ws)


def main():
    payload = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.properties.title = "Saldo final da migração de orçamentos"
    workbook.properties.subject = "Conferência ArkMeds para Ipromed"
    workbook.properties.creator = "Codex"
    workbook.properties.created = datetime(2026, 7, 21)

    create_summary(workbook, payload)
    create_detail_sheet(
        workbook,
        "Cancelados",
        "Orçamentos cancelados ainda não importados",
        "Use a coluna Ação recomendada para separar o primeiro dry-run dos casos que ainda dependem de conferência.",
        payload["cancelados"],
        "TabelaCanceladosSaldoFinal",
    )
    create_detail_sheet(
        workbook,
        "Outros pendentes",
        "Demais orçamentos ainda não importados",
        "Saldo residual fora do grupo cancelado. Recusados permanecem fora da importação até decisão explícita.",
        payload["outros"],
        "TabelaOutrosSaldoFinal",
    )
    create_new_sheet(workbook, payload["novos_sem_staging"])
    create_original_data_sheet(workbook, payload["cancelados"] + payload["outros"])

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
