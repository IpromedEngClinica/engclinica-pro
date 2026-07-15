"""Enriquece o staging de orcamentos ArkMeds a partir da planilha conferida.

Por padrao executa apenas simulacao. Para atualizar o staging use --aplicar.
Nao cria orcamentos nem itens no sistema final.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import psycopg2
from psycopg2.extras import Json, execute_batch
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT_DIR / "outputs" / "migracao-orcamentos"


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_iso_datetime(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()

    text = clean_text(value)
    if not text:
        return None

    for pattern in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).isoformat()
        except ValueError:
            continue
    return text


def to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    text = re.sub(r"[^0-9,.-]", "", text)
    if text.count(",") == 1 and text.count(".") >= 1:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(",") == 1:
        text = text.replace(",", ".")
    try:
        return Decimal(text or "0").quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def read_rows(file_path: Path, limit: int | None) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[1]
    headers = [normalize_header(value) for value in next(worksheet.iter_rows(values_only=True))]
    rows: list[dict[str, Any]] = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        data = {
            headers[index]: json_value(value)
            for index, value in enumerate(values)
            if index < len(headers)
        }
        arkmeds_id = data.get("id")
        try:
            data["arkmeds_orcamento_id"] = int(arkmeds_id)
        except (TypeError, ValueError):
            continue
        rows.append(data)
        if limit and len(rows) >= limit:
            break

    return rows


def build_update_row(row: dict[str, Any]) -> tuple[Any, ...]:
    status_key = normalize_header(row.get("etapa_atual"))
    status_map = {
        "aprovado": ("aprovado_em_curso", "importar_operacional"),
        "cancelado": ("cancelado", "importar_historico"),
        "faturado": ("faturado", "importar_operacional"),
        "pendente": ("pendente", "importar_operacional"),
        "reprovado": ("reprovado_em_curso", "importar_historico"),
    }
    normalized_status, import_policy = status_map.get(status_key, (None, None))
    return (
        to_iso_datetime(row.get("data_de_aprovacao")),
        to_iso_datetime(row.get("data_de_reprovacao")),
        to_iso_datetime(row.get("data_de_faturamento")),
        to_iso_datetime(row.get("data_de_cancelamento")),
        to_decimal(row.get("desconto")),
        "valor",
        to_decimal(row.get("valor_deslocamento")),
        to_decimal(row.get("valor_viagem")),
        to_decimal(row.get("valor_frete")),
        clean_text(row.get("etapa_atual")),
        clean_text(row.get("ordens_de_servico")),
        clean_text(row.get("observacoes")),
        Json(row),
        normalized_status,
        import_policy,
        row["arkmeds_orcamento_id"],
    )


UPDATE_SQL = """
update public.staging_arkmeds_orcamentos
set
  arkmeds_data_aprovacao = %s,
  arkmeds_data_reprovacao = %s,
  arkmeds_data_faturamento = %s,
  arkmeds_data_cancelamento = %s,
  arkmeds_desconto = %s,
  arkmeds_desconto_tipo = %s,
  arkmeds_valor_deslocamento = %s,
  arkmeds_valor_viagem = %s,
  arkmeds_valor_frete = %s,
  arkmeds_status_planilha = %s,
  arkmeds_ordem_servico_planilha = %s,
  arkmeds_observacoes_planilha = %s,
  dados_planilha_json = %s,
  status_normalizado_importacao = coalesce(%s, status_normalizado_importacao),
  politica_importacao_status = coalesce(%s, politica_importacao_status),
  fonte_planilha_atualizada_em = now()
where arkmeds_orcamento_id = %s
"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--arquivo", required=True, type=Path)
    parser.add_argument("--aplicar", action="store_true")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    if not args.arquivo.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {args.arquivo}")
    if not os.environ.get("SUPABASE_DB_URL"):
        raise RuntimeError("Configure SUPABASE_DB_URL antes de executar este script.")

    rows = read_rows(args.arquivo, args.limit)
    ids = [row["arkmeds_orcamento_id"] for row in rows]
    status_counter = Counter(clean_text(row.get("etapa_atual")) or "sem_status" for row in rows)
    desconto_count = sum(1 for row in rows if to_decimal(row.get("desconto")) > 0)

    with psycopg2.connect(os.environ["SUPABASE_DB_URL"], sslmode="require") as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                select arkmeds_orcamento_id
                from public.staging_arkmeds_orcamentos
                where arkmeds_orcamento_id = any(%s)
                """,
                (ids,),
            )
            stage_ids = {row[0] for row in cursor.fetchall()}

            missing_ids = sorted(set(ids) - stage_ids)
            if args.aplicar:
                updates = [build_update_row(row) for row in rows if row["arkmeds_orcamento_id"] in stage_ids]
                execute_batch(cursor, UPDATE_SQL, updates, page_size=250)
                connection.commit()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "arquivo": str(args.arquivo),
        "modo": "aplicar" if args.aplicar else "simulacao",
        "total_planilha": len(rows),
        "encontrados_no_staging": len(stage_ids),
        "nao_encontrados_no_staging": missing_ids,
        "com_desconto": desconto_count,
        "status_planilha": dict(status_counter),
    }
    output_path = OUTPUT_DIR / "enriquecimento_planilha_resultado.json"
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"Relatorio: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
